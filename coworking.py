"""Script de coworking."""

import datetime as dt
import json
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import sqlite3
from sqlite3 import Error
import sys
import os

class Coworking:
    """Clase principal del coworking."""

    def __init__(self):
        if not os.path.exists("coworking.db"):
            print("Aviso: No se encontró la base de datos, por lo que se iniciará con un estado vacío.")
            self.__inicializar_base_datos()

        self.clientes = self.ManejarClientes()
        self.salas = self.ManejarSalas()
        self.reservaciones = self.ManejarReservaciones()

    class ManejarReservaciones:
        """Clase para manejar reservaciones."""

        def __init__(self):
            pass

        def __convertir_turno_a_numero(self, turno: str) -> int:
            """Convierte un string de turno a su número correspondiente.

            Args:
                turno (str): String que representa al turno (Matutino, Vespertino o Nocturno)

            Returns:
                int: Número que representa al turno (1, 2 o 3)
            """

            resultado = 0

            match turno:
                case "Matutino":
                    resultado = 1
                case "Vespertino":
                    resultado = 2
                case "Nocturno":
                    resultado = 3

            return resultado

        def registrar_reservacion(self, id_cliente: int, fecha: dt.date, turno: str, id_sala: int, nombre_evento: str) -> None:
            """Registra una reservación en la base de datos.

            Args:
                id_cliente (int): ID del cliente.
                fecha (dt.date): Fecha de la reservación.
                turno (str): Turno de la reservación.
                id_sala (int): ID de la sala.
                nombre_evento (str): Nombre del evento.
            """

            try:
                fecha_formateada = fecha.isoformat()
                num_turno = self.__convertir_turno_a_numero(turno)
                valores = (id_cliente, fecha_formateada, num_turno, id_sala, nombre_evento)

                with sqlite3.connect("coworking.db") as conn:
                    cursor = conn.cursor()

                    cursor.execute("PRAGMA foreign_keys = ON;")

                    cursor.execute("""
                        INSERT INTO reservaciones (id_cliente, fecha, id_turno, id_sala, nombre_evento)
                        VALUES (?, ?, ?, ?, ?);
                    """, valores)


                print("Evento registrado de manera exitosa.")
            except ValueError as e:
                print(e)
            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

        def mostrar_reservaciones_por_fecha(self, fecha: dt.date, datos: list = None) -> None:
            """Muestra las reservaciones por fecha en formato tabular.

            Args:
                fecha (dt.date): Fecha a consultar.
                datos (list): Lista que contiene los datos de la consulta,
                en caso de haberla hecho previamente. (opcional)
            """

            try:
                if datos:
                    resultados = datos
                else:
                    resultados = self.obtener_reservaciones_por_fecha(fecha)

                if resultados:
                    print(f"\n{'-'*106}")
                    print(f"|{'Folio':^10}|{'Nombre de la sala':^25}|{'Nombre del cliente':^20}|{'Nombre del evento':^30}|{'Turno':^15}|")
                    print('='*106)
                    for folio, nombre_sala, nombre_cliente, nombre_evento, turno in resultados:
                        print(f"|{folio:^10}|{nombre_sala:^25}|{nombre_cliente:^20}|{nombre_evento:^30}|{turno:^15}|")
                    print('-'*106)
                else:
                    print("No hay reservaciones disponibles para esta fecha.")
            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

        def obtener_reservaciones_por_fecha(self, fecha: dt.date) -> list:
            """Obtiene las reservaciones por fecha.

            Args:
                fecha (dt.date): Fecha a consultar.

            Returns:
                list: Lista de tuplas con los datos de las reservaciones.
            """

            fecha_formateada = fecha.isoformat()
            valores = (fecha_formateada,)

            try:
                with sqlite3.connect("coworking.db") as conn:
                    cursor = conn.cursor()

                    cursor.execute("""
                        SELECT
                            r.folio,
                            s.nombre,
                            c.nombre,
                            r.nombre_evento,
                            t.turno
                        FROM reservaciones r
                        JOIN salas s ON s.id_sala = r.id_sala
                        JOIN clientes c ON c.id_cliente = r.id_cliente
                        JOIN turnos t ON r.id_turno = t.id_turno
                        WHERE r.fecha = ?
                        AND r.cancelado IS NULL;
                    """, valores)

                    resultados = cursor.fetchall()

                    return resultados

            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")


        def mostrar_reservaciones_en_rango(self, fecha_inicio: dt.date, fecha_fin: dt.date, datos: list = None) -> list:
            """Muestra las reservaciones como formato tabular dentro de un rango de fechas definido.

            Args:
                fecha_inicio (date): Fecha de inicio a consultar.
                fecha_fin (date): Fecha fin a consultar.
                datos (list): Lista con los datos de las reservaciones. (opcional)


            Returns:
                list: Lista de folios válidos en el rango.
            """

            try:
                if datos:
                    resultados = datos
                else:
                    resultados = self.obtener_reservaciones_en_rango(fecha_inicio, fecha_fin)

                if resultados:
                    print(f"\n{'-'*112}")
                    print(f"|{'Folio':^10}|{'ID cliente':^15}|{'Fecha':^15}|{'Turno':^15}|{'ID sala':^10}|{'Nombre del evento':^40}|")
                    print('='*112)
                    for folio, id_cliente, fecha, turno, id_sala, nombre_evento in resultados:
                        fecha = dt.date.fromisoformat(fecha)
                        fecha = fecha.strftime('%m-%d-%Y')
                        print(f"|{folio:^10}|{id_cliente:^15}|{fecha:^15}|{turno:^15}|{id_sala:^10}|{nombre_evento:^40}|")
                    print("-"*112)
                else:
                    print("No hay reservaciones disponibles para esta fecha.")
            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")


        def obtener_reservaciones_en_rango(self, fecha_inicio: dt.date, fecha_fin: dt.date) -> list:
            """Obtiene las reservaciones especificando un rango de fechas.
            Args:
                fecha_inicio (dt.date): Fecha de inicio a consultar.
                fecha_fin (dt.date): Fecha fin a consultar.
            Returns:
                list: Lista de tuplas con los datos de las reservaciones.
            """

            fecha_inicio = fecha_inicio.isoformat()
            fecha_fin = fecha_fin.isoformat()

            valores = (fecha_inicio, fecha_fin)
            try:
                with sqlite3.connect("coworking.db") as conn:
                    cursor = conn.cursor()

                    cursor.execute("""
                        SELECT
                            r.folio,
                            r.id_cliente,
                            r.fecha,
                            t.turno,
                            r.id_sala,
                            r.nombre_evento
                        FROM reservaciones r
                        JOIN turnos t ON t.id_turno = r.id_turno
                        WHERE fecha BETWEEN ? AND ?
                        AND r.cancelado IS NULL;
                    """, valores)

                    resultados = cursor.fetchall()
                    return resultados
            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

        def editar_nombre_evento(self, folio: int, nuevo_nombre: str) -> None:
            """Edita el nombre de un evento ya existente.

            Args:
                folio (str): Folio del evento.
                nuevo_nombre (str): Nuevo nombre que tendrá el evento.
            """

            valores = (nuevo_nombre, folio)
            try:
                with sqlite3.connect("coworking.db") as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        UPDATE reservaciones
                        SET nombre_evento = ?
                        WHERE folio = ?;
                    """, valores)
                    print("Nombre del evento actualizado exitosamente.")
            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

        def verificar_existencia_reservacion(self, fecha: dt.date, id_sala: int, turno: str) -> bool:
            """Verifica si existe una reservación.

            Args:
                fecha (dt.date): Fecha a consultar.
                id_sala (int): ID de la sala a consultar.
                turno (str): Turno a consultar.

            Returns:
                bool: True si existe, False si no existe.
            """

            fecha_formateada = fecha.isoformat()
            valores = (fecha_formateada, id_sala, turno)

            try:
                with sqlite3.connect("coworking.db") as conn:
                    cursor = conn.cursor()

                    cursor.execute("""
                        SELECT
                            r.id_cliente,
                            r.fecha,
                            t.turno,
                            r.id_sala,
                            r.nombre_evento
                        FROM reservaciones r
                        JOIN turnos t ON t.id_turno = r.id_turno
                        WHERE fecha = ?
                        AND id_sala = ?
                        AND turno = ?
                        AND r.cancelado IS NULL;
                    """, valores)

                    resultados = cursor.fetchall()

                    if resultados:
                        return True
                    else:
                        return False

            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

            return True

        def cancelar_reservación(self, folio: int) -> None:
            """Cancela una reservación, marcándola como cancelada.

            Args:
                folio (int): Folio de la reservación a cancelar.
            """

            valores = (folio,)

            try:
                with sqlite3.connect("coworking.db") as conn:
                    cursor = conn.cursor()

                    cursor.execute("""
                        UPDATE reservaciones
                        SET cancelado = 1
                        WHERE folio = ?;
                    """, valores)

                    print("Reservación cancelada exitosamente.")
            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

    class ManejarSalas:
        """Clase para el manejo de salas."""

        def __init__(self):
            pass

        def registrar_sala(self, nombre: str, cupo: int) -> None:
            """Registra una sala en la base de datos.

            Args:
                nombre (str): Nombre de la sala.
                cupo (int): Cupo de la sala.
            """

            try:
                with sqlite3.connect("coworking.db") as conn:
                    cursor = conn.cursor()

                    cursor.execute("""
                        INSERT INTO salas (nombre, cupo)
                        VALUES (?, ?);
                    """, (nombre, cupo))

                print("Sala registrada exitosamente.")
            except ValueError as e:
                print(e)
            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

        def mostrar_salas_disponibles(self, fecha:dt.date, datos:list = None) -> None:
            """Muestra las salas disponibles en una fecha específica.

            Args:
                lista_salas (dict): Lista que contiene las salas registradas.
                fecha (dt.date): Fecha a consultar.
                datos (list): Lista con los datos de las salas. (opcional)
            """

            try:
                if datos:
                    resultados = datos
                else:
                    resultados = self.obtener_salas_disponibles(fecha)

                if resultados:
                    print(f"\n{'-'*75}")
                    print(f"|{'ID sala':^10}|{'Nombre':^10}|{'Cupo':^10}|{'Turnos disponibles':^40}|")
                    print('='*75)
                    for id_sala, nombre, cupo, turnos_disponibles in resultados:
                        print(f"|{id_sala:^10}|{nombre:^10}|{cupo:^10}|{turnos_disponibles:^40}|")
                    print('-'*75)
                else:
                    print("No hay salas disponibles para esta fecha.")
            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

        def obtener_salas_disponibles(self, fecha:dt.date) -> list:
            """Obtiene las salas disponibles en una fecha específica.

            Args:
                fecha (dt.date): Fecha a consultar.

            Returns:
                list: Lista de tuplas con los datos de las salas.
            """

            fecha_formateada = fecha.isoformat()
            valores = (fecha_formateada,)

            try:
                with sqlite3.connect("coworking.db") as conn:
                    cursor = conn.cursor()

                    cursor.execute("""
                            SELECT
                                s.id_sala,
                                s.nombre,
                                s.cupo,
                                group_concat(t.turno, ', ')
                            FROM salas s
                            CROSS JOIN turnos t
                            LEFT JOIN
                            (
                                SELECT
                                *
                                FROM reservaciones r
                                JOIN salas s ON s.id_sala = r.id_sala
                                WHERE cancelado IS NULL
                                AND r.fecha IS ?
                            ) AS re ON re.id_turno = t.id_turno AND re.id_sala = s.id_sala
                            WHERE re.id_turno IS NULL
                            GROUP BY s.id_sala
                        """, valores)

                    resultados = cursor.fetchall()
                    return resultados

            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

    class ManejarClientes:
        """Clase para el manejo de clientes."""

        def __init__(self):
            pass

        def registrar_cliente(self, nombre: str, apellidos: str) -> None:
            """Registra un cliente en la base de datos.

            Args:
                nombre (str): Nombre del cliente.
                apellidos (str): Apellidos del cliente.
            """

            try:
                with sqlite3.connect("coworking.db") as conn:
                    cursor = conn.cursor()

                    cursor.execute("""
                        INSERT INTO clientes (nombre, apellidos)
                        VALUES (?, ?);
                    """, (nombre, apellidos))
                print("Cliente registrado satisfactoriamente.")
            except ValueError as e:
                print(e)
            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

        def mostrar_clientes(self, datos: list = None) -> None:
            """Muestra los clientes registrados en formato tabular.
            Args:
                datos (list): Lista que contiene los datos de los clientes. (opcional)
            """

            try:
                if datos:
                    resultados = datos
                else:
                    resultados = self.obtener_clientes()

                if resultados:
                    print(f"\n{'-'*63}")
                    print(f"|{'ID':^10}|{'Nombre':^25}|{'Apellidos':^25}|")
                    print(f"\n{'='*63}")
                    for id_cliente, nombre, apellidos in resultados:
                        print(f"|{id_cliente:^10}|{nombre:^25}|{apellidos:^25}|")
                    print('-'*63)
                else:
                    print("No hay clientes registrados.")
            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

        def obtener_clientes(self) -> list:
            """Obtiene los clientes registrados.

            Returns:
                list: Lista de tuplas con los datos de los clientes.
            """

            try:
                with sqlite3.connect("coworking.db") as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                    """
                        SELECT
                            id_cliente,
                            nombre,
                            apellidos
                        FROM clientes
                        ORDER BY apellidos;
                    """
                    )

                    resultados = cursor.fetchall()

                    return resultados

            except Error as e:
                print(e)
            except Exception:
                print(f"Ocurrió un error: {sys.exc_info()[0]}")

    def __inicializar_base_datos(self) -> None:
        """Crea coworking.db y las tablas básicas si no existen."""

        try:
            with sqlite3.connect("coworking.db") as conn:
                cursor = conn.cursor()


                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS clientes (
                        id_cliente INTEGER PRIMARY KEY,
                        nombre TEXT NOT NULL,
                        apellidos TEXT NOT NULL
                    );
                """)


                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS salas (
                        id_sala INTEGER PRIMARY KEY,
                        nombre TEXT NOT NULL,
                        cupo INTEGER NOT NULL
                    );
                """)

                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS turnos (
                        id_turno INTEGER PRIMARY KEY,
                        turno TEXT NOT NULL
                    );
                """)

                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS reservaciones (
                        folio INTEGER PRIMARY KEY,
                        id_cliente INTEGER NOT NULL,
                        fecha TEXT NOT NULL,
                        id_turno INTEGER NOT NULL,
                        id_sala INTEGER NOT NULL,
                        nombre_evento TEXT NOT NULL,
                        cancelado INTEGER,
                        FOREIGN KEY (id_cliente) REFERENCES clientes(id_cliente),
                        FOREIGN KEY (id_sala) REFERENCES salas(id_sala),
                        FOREIGN KEY (id_turno) REFERENCES turnos(id_turno)
                    );
                """)

                cursor.execute("""
                    INSERT INTO turnos (turno) VALUES ('Matutino');
                """)

                cursor.execute("""
                    INSERT INTO turnos (turno) VALUES ('Vespertino');
                """)

                cursor.execute("""
                    INSERT INTO turnos (turno) VALUES ('Nocturno');
                """)

        except Error as e:
            print(e)
        except Exception:
            print(f"Ocurrió un error: {sys.exc_info()[0]}")

    def __verificar_salida(self) -> bool:
        """Verifica si el usuario quiere salir de la operación actual.
        De esta forma evitamos repetir las validaciones flag.

        Returns:
            bool: True si el usuario quiere salir, False si quiere continuar.
        """

        confirmar_salida = input("¿Quiere cancelar la operación? Escriba S para salir o cualquier tecla para continuar: ")
        return True if confirmar_salida.upper() == "S" else False

    def __pedir_string(self, mensaje: str) -> str:
        """Pide un string al usuario y se asegura de que no esté vacío.

        Args:
            mensaje (str): Mensaje a mostrar al usuario.

        Returns:
            str: String introducido por el usuario.
        """

        while True:
            entrada = input(mensaje).strip()

            if not entrada:
                print("El campo no puede estar vacío.")
                raise ValueError

            return entrada

    def __exportar_json(self, reservaciones: dict, fecha: str) -> None:
        """Exporta las reservaciones a formato JSON.

        Args:
            reservaciones (dict): Diccionario con los valores de las reservaciones.
            fecha (str): Fecha de consulta.
        """

        with open(f"reservaciones_{fecha}.json", "w", encoding="utf-8") as archivo:
            json.dump(reservaciones, archivo, indent=2, ensure_ascii=False)
        print(f"Reservaciones exportadas correctamente a 'reservaciones_{fecha}.json'")


    def __exportar_csv(self,reservaciones: dict, fecha: str) -> None:
        """Exporta las reservaciones a formato CSV.

        Args:
            reservaciones (dict): Diccionario con los valores de las reservaciones.
            fecha (str): Fecha de consulta.
        """

        with open(f"reservaciones_{fecha}.csv", "w", newline="", encoding="utf-8") as archivo:
            manejar_csv = csv.writer(archivo)
            manejar_csv.writerow(("Folio", "Nombre Sala", "Nombre Cliente", "Nombre Evento", "Turno", "Fecha"))

            for folio, datos in reservaciones.items():
                manejar_csv.writerow((folio, datos["nombre_sala"], datos["nombre_cliente"], datos["nombre_evento"], datos["turno"], datos["fecha"]))


    def __exportar_excel(self, reservaciones: dict, fecha: str) -> None:
        """Exporta las reservaciones a formato XLSX (Excel)

        Args:
            reservaciones (dict): Diccionario con los valores de las reservaciones.
            fecha (str): Fecha de consulta.
        """

        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = f"Reservaciones_{fecha}"

        negrita = Font(bold=True)
        centrado = Alignment(horizontal="center", vertical="center")
        borde_grueso = Border(bottom=Side(border_style="thick"))

        encabezados = ["Folio", "Nombre Sala", "Nombre Cliente", "Nombre Evento", "Turno", "Fecha"]

        for columna, titulo in enumerate(encabezados, start=1):
            celda = hoja.cell(row=1, column=columna, value=titulo)
            celda.font = negrita
            celda.alignment = centrado
            celda.border = borde_grueso

        for renglon, (folio, datos) in enumerate(reservaciones.items(), start=2):
            hoja.cell(row=renglon, column=1, value=folio).alignment = centrado
            hoja.cell(row=renglon, column=2, value=datos["nombre_sala"]).alignment = centrado
            hoja.cell(row=renglon, column=3, value=datos["nombre_cliente"]).alignment = centrado
            hoja.cell(row=renglon, column=4, value=datos["nombre_evento"]).alignment = centrado
            hoja.cell(row=renglon, column=5, value=datos["turno"]).alignment = centrado
            hoja.cell(row=renglon, column=6, value=datos["fecha"]).alignment = centrado

        libro.save(f"reservaciones_{fecha}.xlsx")
        print(f"Reservaciones exportadas correctamente a 'reservaciones_{fecha}.xlsx'")


    def __exportar(self, lista_reservaciones: list, fecha: dt.date) -> None:
        """Menú que permite elegir el formato a exportar la lista de reservaciones.

        Args:
            lista_reservaciones (list): Lista de tuplas que contiene las reservaciones.
            fecha (dt.date): Fecha de consulta.
        """

        formato = input("Seleccione el formato de exportación: JSON, CSV o EXCEL: ").upper()

        fecha_str = fecha.strftime('%m-%d-%Y')
        reservaciones_fecha = {
            folio[0]: {
                "nombre_sala": folio[1],
                "nombre_cliente": folio[2],
                "nombre_evento": folio[3],
                "turno": folio[4],
                "fecha": fecha_str
                }
            for folio in lista_reservaciones
        }

        if formato == "JSON":
            self.__exportar_json(reservaciones_fecha, fecha_str)

        elif formato == "CSV":
            self.__exportar_csv(reservaciones_fecha, fecha_str)

        elif formato == "EXCEL":
            self.__exportar_excel(reservaciones_fecha, fecha_str)
        else:
            print("Formato no válido. Opciones disponibles: JSON, CSV, EXCEL.")

    def __registrar_reservacion_sala(self) -> None:
        """Opción #1 del menú. Permite registrar la reservación de una sala.

        Returns:
            None: Usado para salir de la función en caso de que el usuario lo decida.
        """

        print("Ha escogido la opción: Registrar reservación de sala")

        lista_clientes = self.clientes.obtener_clientes()
        ids_validos = [cliente[0] for cliente in lista_clientes]

        while True:
            self.clientes.mostrar_clientes(lista_clientes)

            try:
                id_cliente = int(self.__pedir_string("Escriba su ID de cliente: "))

                if id_cliente not in ids_validos:
                    print("ID de cliente no válido.")
                    raise ValueError
            except ValueError:
                if self.__verificar_salida():
                    return
                continue

            break

        while True:
            try:
                fecha_str = self.__pedir_string("Escriba la fecha (mm-dd-yyyy): ")
                fecha = dt.datetime.strptime(fecha_str, "%m-%d-%Y").date()

                fecha_minima = dt.date.today() + dt.timedelta(days=2)

                if fecha < fecha_minima:
                    print("La reservación debe ser por lo menos con dos días de anticipación.")
                    continue

                if fecha.weekday() == 6:
                    print("La fecha solicitada es domingo. No se permiten reservaciones en domingos.")
                    fecha_propuesta = fecha + dt.timedelta(days=1)
                    while fecha_propuesta.weekday() == 6:
                        fecha_propuesta += dt.timedelta(days=1)
                    aceptar = input(f"Se propone la fecha {fecha_propuesta.strftime('%m-%d-%Y')}. ¿Acepta? (S/N): ").upper()
                    if aceptar == 'S':
                        fecha = fecha_propuesta
                    else:
                        continue

                break

            except ValueError:
                print("Formato no válido. Por favor, escríbalo de nuevo usando el formato correcto.")

                if self.__verificar_salida():
                    return

                continue


        lista_salas = self.salas.obtener_salas_disponibles(fecha)
        ids_sala_validos = [sala[0] for sala in lista_salas]

        while True:
            self.salas.mostrar_salas_disponibles(fecha, lista_salas)
            try:
                id_sala = int(self.__pedir_string("Escriba el ID de la sala a escoger: "))
                if id_sala not in ids_sala_validos:
                    print("ID de sala no válido.")
                    raise ValueError
                break
            except ValueError:
                if self.__verificar_salida():
                    return

                continue

        while True:
            turno = input("Escriba el turno a escoger (Matutino, Vespertino, Nocturno): ").capitalize()

            if turno not in ("Matutino", "Vespertino", "Nocturno"):
                print("Turno no válido.")

                if self.__verificar_salida():
                    return

                continue

            reservado = self.reservaciones.verificar_existencia_reservacion(fecha, id_sala, turno)

            if reservado:
                print("No hay disponibilidad en ese turno para esta sala.")
                if self.__verificar_salida():
                    return

                continue

            break

        print("Hay disponibilidad.")

        while True:
            try:
                nombre_evento = self.__pedir_string("Escriba el nombre del evento: ")
                break
            except ValueError:
                if self.__verificar_salida():
                    return
                continue

        self.reservaciones.registrar_reservacion(id_cliente, fecha, turno, id_sala, nombre_evento)

    def __editar_nombre_reservacion(self) -> None:
        """Opción #2 del menú. Permite editar el nombre de una reservación ya hecha.

        Returns:
            None: Usado para salir de la función en caso de que el usuario lo decida.
        """

        print("Ha escogido la opción: Editar el nombre de una reservación ya hecha")

        while True:
            try:
                fecha_inicio_str = input("Escriba la fecha de inicio del rango (mm-dd-yyyy): ")
                fecha_inicio = dt.datetime.strptime(fecha_inicio_str, "%m-%d-%Y").date()

                fecha_fin_str = input("Escriba la fecha de fin del rango (mm-dd-yyyy): ")
                fecha_fin = dt.datetime.strptime(fecha_fin_str, "%m-%d-%Y").date()

                if fecha_inicio > fecha_fin:
                    print("La fecha de inicio no puede ser posterior a la de fin.")
                    continue
                break
            except ValueError:
                print("Formato no válido. Por favor, escríbalo de nuevo usando el formato correcto.")

                if self.__verificar_salida():
                    return

                continue

        reservaciones = self.reservaciones.obtener_reservaciones_en_rango(fecha_inicio, fecha_fin)
        folios_validos = [folio[0] for folio in reservaciones]

        while True:
            try:
                self.reservaciones.mostrar_reservaciones_en_rango(fecha_inicio, fecha_fin, reservaciones)
                folio = int(self.__pedir_string("Escriba el folio del evento a modificar: "))

                if folio not in folios_validos:
                    print("Folio no válido en este rango. Por favor, seleccione uno de la lista.")
                    raise ValueError

                break
            except ValueError:
                if self.__verificar_salida():
                    return
                continue


        while True:
            try:
                nuevo_nombre = self.__pedir_string("Escriba el nuevo nombre del evento: ")
                break
            except ValueError:
                if self.__verificar_salida():
                    return
                continue

        self.reservaciones.editar_nombre_evento(folio, nuevo_nombre)

    def __consultar_reservaciones_fecha(self) -> None:
        """Opción #3 del menú. Permite consultar las reservaciones de una fecha específica.

        Returns:
            None: Usado para salir de la función en caso de que el usuario lo decida.
        """

        print("Ha escogido la opción: Consultar reservaciones de una fecha específica")

        while True:
            try:
                fecha_str = input("Escriba la fecha a consultar (mm-dd-yyyy) o presione Enter para la fecha actual: ").strip()

                if not fecha_str:
                    fecha_str = dt.date.today().strftime("%m-%d-%Y")

                fecha = dt.datetime.strptime(fecha_str, "%m-%d-%Y").date()

                break
            except ValueError:
                print("Formato no válido. Por favor, escríbalo de nuevo usando el formato correcto.")

                if self.__verificar_salida():
                    return
                continue

        registros_encontrados = self.reservaciones.obtener_reservaciones_por_fecha(fecha)

        if registros_encontrados:
            self.reservaciones.mostrar_reservaciones_por_fecha(fecha, registros_encontrados)
            exportar =  input("¿Desea exportar estas reservaciones? (SI/NO): ").upper()
            if exportar == "SI":
                self.__exportar(registros_encontrados, fecha)
        else:
            print("No se encontraron reservaciones para el día especificado.")


    def __cancelar_reservacion(self) -> None:
        """Opción #4 del menú. Permite cancelar una reservación.

        Returns:
            None: Usado para salir de la función en caso de que el usuario lo decida.
        """

        print("Para cancelar su reservación, primero especifique el rango de fechas")

        while True:
            try:
                fecha_inicio_str = input("Escriba la fecha de inicio del rango (mm-dd-yyyy): ")
                fecha_inicio = dt.datetime.strptime(fecha_inicio_str, "%m-%d-%Y").date()

                fecha_fin_str = input("Escriba la fecha de fin del rango (mm-dd-yyyy): ")
                fecha_fin = dt.datetime.strptime(fecha_fin_str, "%m-%d-%Y").date()

                if fecha_inicio > fecha_fin:
                    print("La fecha de inicio no puede ser posterior a la de fin.")
                    continue
                break
            except ValueError:
                print("Formato no válido. Por favor, escríbalo de nuevo usando el formato correcto.")

                if self.__verificar_salida():
                    return

                continue


        resultados = self.reservaciones.obtener_reservaciones_en_rango(fecha_inicio, fecha_fin)
        self.reservaciones.mostrar_reservaciones_en_rango(fecha_inicio, fecha_fin, resultados)
        folios_validos = [folio[0] for folio in resultados]

        while True:
            try:
                folio = int(self.__pedir_string("Escriba el folio de la reservación a cancelar: "))

                if folio not in folios_validos:
                    print("Folio no válido.")
                    raise ValueError

                break
            except ValueError:
                if self.__verificar_salida():
                    return
                continue

        print(f"¿Realmente quiere cancelar esta reservación con el folio {folio}?")

        while True:
            eliminar = self.__pedir_string("(S) - Sí, (N) - No: ").upper()

            if eliminar == "S":
                self.reservaciones.cancelar_reservación(folio)
                break
            elif eliminar == "N":
                return



    def __registrar_nuevo_cliente(self) -> None:
        """Opción #5 del menú. Permite registrar a un nuevo cliente.

        Returns:
            None: Usado para salir de la función en caso de que el usuario lo decida.
        """

        print("Ha escogido la opción: Registrar a un nuevo cliente.")

        while True:
            try:
                nombre = self.__pedir_string("Escriba su nombre: ")
                break
            except ValueError:
                if self.__verificar_salida():
                    return
                continue

        while True:
            try:
                apellidos = self.__pedir_string("Escriba sus apellidos: ")
                break
            except ValueError:
                if self.__verificar_salida():
                    return
                continue

        self.clientes.registrar_cliente(nombre, apellidos)

    def __registrar_nueva_sala(self) -> None:
        """Opción #6 del menú. Registra una nueva sala.

        Returns:
            None: Usado para salir de la función en caso de que el usuario lo decida.
        """

        print("Ha escogido la opción: registrar nueva sala")

        while True:
            try:
                nombre_sala = self.__pedir_string("Escriba el nombre de la sala: ")
                break
            except ValueError:
                if self.__verificar_salida():
                    return
                continue

        while True:
            try:
                cupo_str = self.__pedir_string("Escriba el cupo de la sala: ")
                cupo = int(cupo_str)

                if cupo <= 0:
                    print("El cupo debe ser mayor a cero.")
                    raise ValueError
                break

            except ValueError:
                print("Error: Valor inválido")
                if self.__verificar_salida():
                    return
                continue

        self.salas.registrar_sala(nombre_sala, cupo)

    def mostrar_menu(self) -> None:
        """Muestra al usuario una interfaz de texto para poder realizar diversas acciones dentro del coworking."""

        while True:
            print("\nBienvenido al programa del coworking.")
            print("Opciones disponibles:")
            print("(1) - Registrar reservación de sala")
            print("(2) - Editar el nombre de una reservación ya hecha")
            print("(3) - Consultar reservaciones de una fecha específica")
            print("(4) - Cancelar una reservación")
            print("(5) - Registrar a un nuevo cliente")
            print("(6) - Registrar una sala")
            print("(7) - Salir del programa\n")

            while True:
                try:
                    opcion = int(input("Escribe el número de la opción que vas a escoger: "))

                    if opcion < 1 or opcion > 7:
                        print("ERROR: Opción no válida. Escoge entre 1 y 6.")
                        continue

                except ValueError:
                    print("ERROR: Escribe un número válido.")
                    continue

                else:
                    break

            match opcion:
                case 1:
                    self.__registrar_reservacion_sala()
                case 2:
                    self.__editar_nombre_reservacion()
                case 3:
                    self.__consultar_reservaciones_fecha()
                case 4:
                    self.__cancelar_reservacion()
                case 5:
                    self.__registrar_nuevo_cliente()
                case 6:
                    self.__registrar_nueva_sala()
                case 7:
                    confirmar = input("¿Desea salir del programa, los datos se guardaran en la base de datos? (S/N): ").upper()
                    if confirmar == "S":
                        print("Saliendo del programa...")
                        break

if __name__ == "__main__":
    programa = Coworking()
    programa.mostrar_menu()
